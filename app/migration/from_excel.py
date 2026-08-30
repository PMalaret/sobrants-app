"""Migració de dades des de SobrantsV4.74.xlsm a la base de dades SQLite nova.

Llegeix les fulles Materials, llista i històric (Entrades és una còpia redundant
de llista, s'omet) i desmagatzem, i les bolca a l'esquema relacional nou,
preservant l'estat exacte de l'inventari en el moment de la migració.

Ús:
    python -m app.migration.from_excel <ruta_excel.xlsm> <ruta_desti.db>
"""
from __future__ import annotations

import datetime as dt
import sys
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).resolve().parents[2]))
from app.data.db import connect  # noqa: E402

EMPTY_MARK = "---------"


def _norm_dt(value) -> str | None:
    if value is None or value == "":
        return None
    if isinstance(value, dt.datetime):
        return value.isoformat(sep=" ")
    if isinstance(value, dt.date):
        return dt.datetime.combine(value, dt.time()).isoformat(sep=" ")
    return str(value)


def migrate_materials(wb, conn) -> int:
    ws = wb["Materials"]
    rows = []
    for r in range(2, ws.max_row + 1):
        code = ws.cell(row=r, column=1).value
        desc = ws.cell(row=r, column=2).value
        if code is None or code == "":
            continue
        rows.append((int(code), str(desc) if desc is not None else ""))
    conn.executemany(
        "INSERT OR REPLACE INTO materials(code, description) VALUES (?, ?)", rows
    )
    conn.commit()
    return len(rows)


def migrate_pieces(wb, conn) -> int:
    ws = wb["llista"]
    count = 0
    # Cada posició ocupa 5 files consecutives (slots 1..5), ja ordenades per
    # posició a l'Excel original (es reordena en cada obertura/tancament).
    slot_counters: dict[int, int] = {}
    for r in range(1, ws.max_row + 1):
        position = ws.cell(row=r, column=1).value
        if position in (None, ""):
            continue
        position = int(position)
        code = ws.cell(row=r, column=2).value
        desc = ws.cell(row=r, column=3).value
        dims = ws.cell(row=r, column=4).value
        notes = ws.cell(row=r, column=5).value
        entered_at = _norm_dt(ws.cell(row=r, column=6).value)

        slot = slot_counters.get(position, 0) + 1
        slot_counters[position] = slot
        if slot > 5:
            continue  # per seguretat; no hauria de passar en dades vàlides

        is_empty = code in (None, "") and (desc in (None, "", EMPTY_MARK))
        if is_empty:
            continue

        conn.execute(
            """INSERT OR REPLACE INTO pieces
               (position, slot, material_code, material_desc, dimensions, notes, entered_at)
               VALUES (?, ?, ?, ?, ?, ?, ?)""",
            (
                position,
                slot,
                int(code) if isinstance(code, (int, float)) else None,
                str(desc) if desc not in (None, "") else None,
                str(dims) if dims not in (None, "") else None,
                str(notes) if notes not in (None, "") else None,
                entered_at,
            ),
        )
        count += 1
    conn.commit()
    return count


def _parse_kind(e_value):
    """Tradueix la columna E de històric (1 / -1 / fletxes de moviment)."""
    if isinstance(e_value, (int, float)):
        if e_value > 0:
            return 1, "in"
        if e_value < 0:
            return -1, "out"
    text = str(e_value)
    if "⇒" in text:  # ⇒⇒ origen d'un moviment
        return None, "move_out"
    if "⇐" in text:  # ⇐⇐ destí d'un moviment
        return None, "move_in"
    return None, "in"  # valor inesperat; es conserva sense perdre el registre


def migrate_historic(wb, conn) -> int:
    ws = wb["històric"]
    count = 0
    for r in range(2, ws.max_row + 1):
        position = ws.cell(row=r, column=1).value
        if position in (None, ""):
            continue
        material_code = ws.cell(row=r, column=2).value
        material_desc = ws.cell(row=r, column=3).value
        ts = _norm_dt(ws.cell(row=r, column=4).value)
        e_value = ws.cell(row=r, column=5).value
        direction, kind = _parse_kind(e_value)

        conn.execute(
            """INSERT INTO historic(position, material_code, material_desc, ts, direction, kind)
               VALUES (?, ?, ?, ?, ?, ?)""",
            (
                str(position),
                str(material_code) if material_code not in (None, "") else None,
                str(material_desc) if material_desc not in (None, "") else None,
                ts or dt.datetime.now().isoformat(sep=" "),
                direction,
                kind,
            ),
        )
        count += 1
    conn.commit()
    return count


def migrate_desmagatzem(wb, conn) -> int:
    ws = wb["desmagatzem"]
    count = 0
    for order, r in enumerate(range(2, ws.max_row + 1), start=1):
        material_code = ws.cell(row=r, column=2).value
        if material_code in (None, ""):
            continue
        quantity = ws.cell(row=r, column=1).value or 0
        material_desc = ws.cell(row=r, column=3).value
        dimensions = ws.cell(row=r, column=4).value
        cart_ref = ws.cell(row=r, column=5).value
        ts = _norm_dt(ws.cell(row=r, column=6).value)

        code_str = str(material_code)
        custom_text = str(material_desc) if code_str == "1" and material_desc else None

        conn.execute(
            """INSERT INTO desmagatzem
               (row_order, quantity, material_code, material_desc, custom_text, dimensions, cart_ref, ts)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?)""",
            (
                order,
                int(quantity) if isinstance(quantity, (int, float)) else 0,
                code_str,
                str(material_desc) if material_desc not in (None, "") else None,
                custom_text,
                str(dimensions) if dimensions not in (None, "") else None,
                str(cart_ref) if cart_ref not in (None, "") else None,
                ts,
            ),
        )
        count += 1
    conn.commit()
    return count


def migrate(excel_path: str, db_path: str) -> dict:
    wb = openpyxl.load_workbook(excel_path, data_only=True, keep_vba=False)
    # Cridat des de la línia d'ordres, la carpeta de destí (SobrantsData/)
    # encara pot no existir: en el flux normal la crea `main._data_dir` en
    # arrencar l'aplicació, però aquí no hi hem passat i sqlite3 no la crea
    # sol (fallava amb "unable to open database file").
    Path(db_path).parent.mkdir(parents=True, exist_ok=True)
    conn = connect(db_path)
    try:
        stats = {
            "materials": migrate_materials(wb, conn),
            "pieces": migrate_pieces(wb, conn),
            "historic": migrate_historic(wb, conn),
            "desmagatzem": migrate_desmagatzem(wb, conn),
        }
        return stats
    finally:
        conn.close()


if __name__ == "__main__":
    if len(sys.argv) != 3:
        print("Ús: python -m app.migration.from_excel <origen.xlsm> <desti.db>")
        raise SystemExit(1)
    result = migrate(sys.argv[1], sys.argv[2])
    print("Migració completada:")
    for k, v in result.items():
        print(f"  {k}: {v} files")
